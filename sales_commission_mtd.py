"""
Sales Commission — MTD (per salesperson)
Standalone Streamlit app (Community Cloud, like revops_signal_dashboard.py):
shared-password gate + key-pair Snowflake service account.

Purpose: month-end commission worksheet. Shows each rep's MTD revenue + order
count (validated against the "sales-by-salesperson" export to the dollar), their
MTD target, and a $1,500 prepayment column for Inside Sales. Comm % and Payout
are left blank in-app — HR fills Comm % at month-end. The "Export to Excel"
download ships a workbook with LIVE FORMULAS so typing a % into Comm %
auto-computes Payout (= Revenue × Comm%) and Net Payout (= Payout − Prepayment).

Revenue math (reverse-engineered + validated 2026-06-17):
  grain   = SALES_ORDER_SHIPMENT_LINES
  revenue = Σ(shipment-line QUANTITY × order-line UNIT_PRICE)
  period  = shipment COMPLETED_AT in the current calendar month
  #orders = distinct SO_NUMBER
  Sales Admin bucket excluded (moving in-house sales to the group soon).

Manual trackers (added 2026-06-22, values entered by hand in the Excel export
while the app stays read-only):
  • Inside Sales draw balance — cumulative advance vs. commission recovered to
    date (separate from the recurring $1,500/mo prepayment, which is unchanged).
  • Door coverage — active retail locations now vs. end-of-quarter target per
    field rep (leading indicator for next quarter), plus Corey's national-account
    doors tracked in a separate block.
"""
import calendar
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date
from snowflake.snowpark.context import get_active_session

st.set_page_config(page_title="Sales Commission — MTD", layout="wide")

# ── Roster ──────────────────────────────────────────────────────────────────
# Revenue/orders now come from GOLD_V3_DB.SCORECARD.FCT_SALES_REP_REVENUE_DAILY,
# which already carries REP_DISPLAY_NAME + ROLE + IS_SCORECARD_REP (verified equal
# to the prior validated shipment-completion math, 2026-06-18). So no more ADP name
# crosswalk / Gonzales-vs-Gonzalez handling — we join by display name. This roster
# stays only to (a) render reps with $0 MTD, (b) hold targets + the prepayment rule.
# mtd_target: derived from the Dec-2027 monthly EXIT targets discounted at each rep's
# CMGR to the June-2026 anchor (~$500K). PENDING RICHIE'S CONFIRMATION. William
# (Inside Sales) + Corey (VP) have no exit target → None (blank).
ROSTER = [
    # display,            role,                              inside_sales, mtd_target
    ("Miguel Gonzalez",   "Regional Sales Manager",          False, 502_000),
    ("Melinda Kingston",  "Regional Sales Manager",          False, 502_000),
    ("Santo Perry",       "Regional Sales Manager",          False, 497_000),
    ("Nelson Rosario",    "Regional Sales Manager",          False, 500_000),
    ("Kamala Watkins",    "Key Account Manager",             False, 496_000),
    ("Quinn McHenry",     "Regional Sales Manager",          False, 496_000),
    ("William Stevens",   "Inside Sales Representative",      True,  None),
    ("Corey Helper",      "VP, Sales & Commercial Strategy",  False, None),
]
INSIDE_SALES_PREPAYMENT = 1_500.0

# ── Shared-password gate (mirrors revops_signal_dashboard.py) ───────────────
def _check_password():
    try:
        if "APP_PASSWORD" not in st.secrets:
            return True  # no APP_PASSWORD → Streamlit-in-Snowflake (ambient auth)
    except Exception:
        return True  # no secrets.toml at all → don't gate (ambient/unconfigured)
    if st.session_state.get("_authed"):
        return True
    st.title("Sales Commission — MTD")
    pw = st.text_input("Password", type="password", label_visibility="collapsed",
                       placeholder="Password")
    if pw and pw == st.secrets["APP_PASSWORD"]:
        st.session_state["_authed"] = True
        st.rerun()
    elif pw:
        st.error("Incorrect password.")
    return False

if not _check_password():
    st.stop()

# ── Snowflake session (ambient inside SF; key-pair service acct on Cloud) ───
@st.cache_resource(show_spinner=False)
def _get_session():
    try:
        return get_active_session()
    except Exception:
        pass
    from snowflake.snowpark import Session
    from cryptography.hazmat.primitives import serialization
    pk = serialization.load_pem_private_key(
        st.secrets["sf_private_key"].encode(),
        password=(st.secrets["sf_private_key_passphrase"].encode()
                  if st.secrets.get("sf_private_key_passphrase") else None),
    ).private_bytes(
        encoding=serialization.Encoding.DER,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=serialization.NoEncryption(),
    )
    return Session.builder.configs({
        "account": st.secrets["sf_account"],
        "user": st.secrets["sf_user"],
        "private_key": pk,
        "role": st.secrets.get("sf_role", "REVOPS_DASHBOARD_RO"),
        "warehouse": st.secrets["sf_warehouse"],
        "database": st.secrets.get("sf_database", "GOLD_V3_DB"),
        "schema": st.secrets.get("sf_schema", "PUBLIC"),
        # Heartbeat keeps the token fresh while the app is awake. On Community
        # Cloud the app sleeps and the heartbeat freezes, so the cached session
        # can still go stale — q() handles that by rebuilding on token expiry.
        "client_session_keep_alive": True,
    }).create()

session = _get_session()

# Snowflake auth-token-expired error codes. The cached @st.cache_resource
# session outlives the token (esp. across Community Cloud sleep/wake), so when
# we see one of these we drop the dead session, rebuild, and retry once.
_EXPIRED_TOKEN_CODES = ("390114", "390108", "08001")

@st.cache_data(ttl=600, show_spinner=False)
def q(sql):
    global session
    try:
        return session.sql(sql).to_pandas()
    except Exception as e:
        if any(code in str(e) for code in _EXPIRED_TOKEN_CODES):
            _get_session.clear()       # evict the dead shared session
            session = _get_session()   # rebuild with a fresh token
            return session.sql(sql).to_pandas()
        raise

# ── Data: MTD revenue + orders per rep, from the dbt scorecard model ─────────
# GOLD_V3_DB.SCORECARD.FCT_SALES_REP_REVENUE_DAILY (REVENUE_DATE = shipment
# COMPLETED_AT; Sales Admin / DTC excluded via IS_SCORECARD_REP). Verified equal to
# the prior V_SALES_COMMISSION_MTD math to the dollar, so those interim views are
# retired. REVOPS_DASHBOARD_RO has SELECT + schema USAGE on SCORECARD.
@st.cache_data(ttl=600, show_spinner=False)
def load_mtd():
    return q("""
        SELECT REP_DISPLAY_NAME AS REP,
               ROUND(SUM(TOTAL_REVENUE_AMT),2) AS MTD_REVENUE,
               SUM(ORDER_COUNT)                AS NUM_ORDERS
        FROM GOLD_V3_DB.SCORECARD.FCT_SALES_REP_REVENUE_DAILY
        WHERE IS_SCORECARD_REP = TRUE
          AND REVENUE_DATE >= DATE_TRUNC('month', CURRENT_DATE())
        GROUP BY REP_DISPLAY_NAME
    """)

mtd = load_mtd()
rev_by_rep = dict(zip(mtd["REP"], mtd["MTD_REVENUE"]))
ord_by_rep = dict(zip(mtd["REP"], mtd["NUM_ORDERS"]))

rows = []
for display, role, inside, target in ROSTER:
    rows.append({
        "Rep": display,
        "Role": role,
        "MTD Revenue": float(rev_by_rep.get(display, 0) or 0),
        "# Orders": int(ord_by_rep.get(display, 0) or 0),
        "MTD Target": target,                                   # may be None
        "Comm %": None,                                         # HR fills at month-end
        "Payout": None,                                         # = Revenue × Comm%
        "Prepayment": INSIDE_SALES_PREPAYMENT if inside else 0.0,
    })
df = pd.DataFrame(rows).sort_values("MTD Revenue", ascending=False).reset_index(drop=True)

# ── Display ─────────────────────────────────────────────────────────────────
st.title("Sales Commission — MTD")
st.caption(f"Month-to-date through {date.today():%b %-d, %Y} · revenue recognized on shipment "
           f"completion · Sales Admin excluded · Comm % / Payout filled by HR at month-end")

show = df.copy()
show["MTD Revenue"] = show["MTD Revenue"].map(lambda v: f"${v:,.0f}")
show["MTD Target"]  = show["MTD Target"].map(lambda v: f"${v:,.0f}" if pd.notna(v) else "—")
show["Comm %"]      = "—"
show["Payout"]      = "—"
show["Prepayment"]  = show["Prepayment"].map(lambda v: f"${v:,.0f}" if v else "—")
st.dataframe(show, use_container_width=True, hide_index=True)

NATIONAL_ROLE = "VP, Sales & Commercial Strategy"

# end of the current calendar quarter (for the door target horizon)
_q_end_month = ((date.today().month - 1) // 3) * 3 + 3
quarter_end = date(date.today().year, _q_end_month,
                   calendar.monthrange(date.today().year, _q_end_month)[1])

# ── Inside Sales draw balance (running advance vs. recovery) ─────────────────
# The monthly $1,500 prepayment in the table above is unchanged; this tracks the
# CUMULATIVE advance and how much commission has recovered it over time. Advanced /
# Recovered are entered by hand in the Excel export (app is read-only).
st.subheader("Inside Sales — Draw Balance")
st.caption("Cumulative advance vs. commission recovered to date. The monthly $1,500 "
           "prepayment above still applies; enter Recovered to Date in the Excel export.")
draw_rows = [{
    "Rep": display,
    "Advanced to Date": f"${INSIDE_SALES_PREPAYMENT:,.0f}",
    "Recovered to Date": "—",
    "Remaining Balance": "—",
} for display, role, inside, target in ROSTER if inside]
st.dataframe(pd.DataFrame(draw_rows), use_container_width=True, hide_index=True)

# ── Door coverage (leading indicator for next quarter's revenue) ────────────
st.subheader("Door Coverage — Retail")
st.caption(f"Active retail locations each rep is selling into now vs. their target by "
           f"end of quarter ({quarter_end:%b %-d, %Y}). Enter Active / Target in the Excel export.")
door_rows = [{
    "Rep": display,
    "Active Doors (now)": "—",
    "EOQ Target": "—",
    "Doors to Go": "—",
} for display, role, inside, target in ROSTER if role != NATIONAL_ROLE]
st.dataframe(pd.DataFrame(door_rows), use_container_width=True, hide_index=True)

st.subheader("National Account Doors — Corey Helper")
st.caption("National retail locations delivered vs. target, tracked separately from the "
           "field reps' retail doors. Enter per-account numbers in the Excel export.")
st.dataframe(pd.DataFrame([{
    "Doors Delivered": "—", "Target": "—", "Doors to Go": "—",
}]), use_container_width=True, hide_index=True)

# ── Excel export with LIVE formulas ─────────────────────────────────────────
def build_excel(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "MTD Commission"
    headers = ["Rep", "Role", "MTD Revenue", "# Orders", "MTD Target",
               "Comm %", "Payout", "Prepayment", "Net Payout",
               "Active Doors", "EOQ Target", "Doors to Go",
               "Advanced to Date", "Recovered to Date", "Remaining Balance"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="2C5784")
    hdr_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDE2E9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center"); cell.border = border

    for i, rec in enumerate(df.to_dict("records"), start=2):
        ws.cell(i, 1, rec["Rep"])
        ws.cell(i, 2, rec["Role"])
        ws.cell(i, 3, round(float(rec["MTD Revenue"]), 2))           # MTD Revenue
        ws.cell(i, 4, int(rec["# Orders"]))                          # # Orders
        if pd.notna(rec["MTD Target"]):
            ws.cell(i, 5, float(rec["MTD Target"]))                  # MTD Target
        ws.cell(i, 6, None)                                          # Comm % — HR input
        ws.cell(i, 7, f"=C{i}*F{i}")                                 # Payout = Revenue × Comm%
        ws.cell(i, 8, float(rec["Prepayment"]))                      # Prepayment
        ws.cell(i, 9, f"=G{i}-H{i}")                                 # Net Payout = Payout − Prepayment
        # Door coverage (all reps) — Active / EOQ Target filled by hand
        ws.cell(i, 10, None)                                         # Active Doors (now)
        ws.cell(i, 11, None)                                         # EOQ Target
        ws.cell(i, 12, f"=K{i}-J{i}")                                # Doors to Go = Target − Active
        # Inside Sales draw balance (only reps carrying an advance/prepayment)
        if float(rec["Prepayment"]) > 0:
            ws.cell(i, 13, float(INSIDE_SALES_PREPAYMENT))           # Advanced to Date — edit if more advanced
            ws.cell(i, 14, None)                                     # Recovered to Date — fill in by hand
            ws.cell(i, 15, f"=M{i}-N{i}")                            # Remaining Balance = Advanced − Recovered
        for col in ["C", "E", "G", "H", "I", "M", "N", "O"]:
            ws[f"{col}{i}"].number_format = '$#,##0.00'
        ws[f"F{i}"].number_format = '0.0%'                           # type 5% (or 0.05)
        for col in ["J", "K", "L"]:
            ws[f"{col}{i}"].number_format = '0'                      # door counts (whole numbers)
        for c in range(1, len(headers) + 1):
            ws.cell(i, c).border = border

    widths = {"A": 20, "B": 30, "C": 15, "D": 9, "E": 14, "F": 9, "G": 15,
              "H": 13, "I": 15, "J": 12, "K": 11, "L": 11,
              "M": 16, "N": 17, "O": 17}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # note row explaining the input cells + formulas (now all on one sheet)
    note = ws.cell(len(df) + 3, 1,
                   f"Fill by hand: Comm % (e.g. 5%); Active Doors + EOQ Target for every rep "
                   f"(EOQ = quarter end {quarter_end:%b %-d, %Y}); Recovered to Date for Inside Sales. "
                   f"Live formulas: Payout = Revenue × Comm% · Net Payout = Payout − Prepayment · "
                   f"Doors to Go = EOQ Target − Active Doors · Remaining Balance = Advanced − Recovered "
                   f"(Inside Sales draw; negative = owed back). Corey's door columns = national-account doors.")
    note.font = Font(italic=True, color="6B7280")

    bio = BytesIO(); wb.save(bio); return bio.getvalue()

st.download_button(
    "⬇ Export to Excel (with live commission formulas)",
    data=build_excel(df),
    file_name=f"mtd_commission_{date.today():%Y-%m}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
